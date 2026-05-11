"""
Step 1: Parse human-labelled xlsx into a clean benchmark dataset.

Input:  judge/IO/step1_parse/human_labels.xlsx  (downloaded from Google Sheet
        via `?format=xlsx` so cell background colors are preserved)
Output: judge/IO/step1_parse/human_labels.jsonl

Only the active (first) sheet is parsed — the Crocs HR Manager JD.
The workbook contains 18 tabs (one per JD); expanding to all tabs is in
BACKLOG.md.

Row flag comes from the cell background color. Semantics (from user):
    red       -> MISMATCH   (FFFF0000, FFCC0000)  — not a fit for this search
    yellow    -> OK         (FFFFFF00)            — acceptable candidate
    green     -> RELEVANT   (FF00FF00)            — strong match, prize pick
    unflagged -> undecided  (no fill)             — not yet reviewed
"""

import json
from pathlib import Path
from collections import Counter

import openpyxl

HERE = Path(__file__).parent
XLSX = HERE / "IO" / "step1_parse" / "human_labels.xlsx"
OUT = HERE / "IO" / "step1_parse" / "human_labels.jsonl"

COLOR_TO_FLAG = {
    "FFFF0000": "red",
    "FFCC0000": "red",
    "FFFFFF00": "yellow",
    "FF00FF00": "green",
}


def cell_flag(cell) -> str:
    fill = cell.fill
    if not fill or not fill.fgColor or fill.fgColor.type != "rgb":
        return "unflagged"
    return COLOR_TO_FLAG.get(fill.fgColor.rgb, "unflagged")


def split_list(cell_value) -> list[str]:
    if not cell_value:
        return []
    return [s.strip() for s in str(cell_value).split(",") if s.strip()]


def main() -> None:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb.active

    parsed = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name:
            continue
        parsed.append({
            "sheet": ws.title,
            "row": r,
            "name": str(name).strip(),
            "linkedin_url": (ws.cell(r, 2).value or "").strip(),
            "human_score": ws.cell(r, 3).value,
            "email": (ws.cell(r, 4).value or "").strip(),
            "strengths": split_list(ws.cell(r, 5).value),
            "weaknesses": split_list(ws.cell(r, 6).value),
            "missing_data": split_list(ws.cell(r, 7).value),
            "actionable_insights": split_list(ws.cell(r, 8).value),
            "remark_col_i": (str(ws.cell(r, 9).value).strip() if ws.cell(r, 9).value else ""),
            "remark": (str(ws.cell(r, 10).value).strip() if ws.cell(r, 10).value else ""),
            "human_flag": cell_flag(ws.cell(r, 1)),
        })

    with OUT.open("w") as f:
        for item in parsed:
            f.write(json.dumps(item, ensure_ascii=False) + "\n")

    flag_counts = Counter(p["human_flag"] for p in parsed)
    with_remark = sum(1 for p in parsed if p["remark"])
    print(f"Parsed {len(parsed)} candidates from sheet '{ws.title}' -> {OUT}")
    print(f"Flag distribution (real colors): {dict(flag_counts)}")
    print(f"Rows with remarks: {with_remark}")


if __name__ == "__main__":
    main()
