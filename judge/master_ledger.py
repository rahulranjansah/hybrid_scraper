"""
Master ledger — cross-run dedupe for the Crocs pipeline.

Two xlsx files at ``judge/IO/master/``:

  untagged.xlsx   — every candidate we've ever SEEN (no judge rating
                     required). The dedupe source of truth. Scraper
                     results are matched against this before the judge
                     runs, so we never pay to re-judge a known person.
  predictions.xlsx — every candidate the judge has RATED. The audit
                     trail of label + reasoning. A superset-relation
                     to untagged: every row in predictions is also in
                     untagged (plus their judge outputs).

Primary key for dedupe: normalised LinkedIn URL; secondary: lowercase name.

Seed order: Crocs sheet (174) -> step5a preds (14) -> step5c preds (19).
Each new pipeline run appends any NEW names to untagged, and any NEW
rated candidates to predictions.

This module is the single place that reads/writes the ledger. All other
code (run_crocs_hr.py, step5a/step5c) should go through these helpers.
"""

from __future__ import annotations

import datetime as dt
import json
from pathlib import Path
from typing import Iterable

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

HERE = Path(__file__).parent
MASTER_DIR = HERE / "IO" / "master"
UNTAGGED = MASTER_DIR / "untagged.xlsx"
PREDICTIONS = MASTER_DIR / "predictions.xlsx"
UNTAGGED_JSON = MASTER_DIR / "untagged.json"
PREDICTIONS_JSON = MASTER_DIR / "predictions.json"

# The Crocs sheet is the canonical source of truth for people already
# in the client's system. We do NOT duplicate those rows into the master
# xlsx files — the master files contain only NET-NEW candidates
# discovered by the pipeline. Crocs names are still used for dedupe via
# this jsonl, loaded by `exclusion_keys()`.
CROCS_HUMAN_LABELS = HERE / "IO" / "step1_parse" / "human_labels.jsonl"

UNTAGGED_HEADERS = [
    "name", "linkedin_url", "email", "first_seen_source", "first_seen_at",
]
PREDICTIONS_HEADERS = [
    "name", "linkedin_url", "email",
    "predicted_label", "reasoning_tags", "reasoning_text",
    "red_bucket", "reapproach_after",
    "rated_at", "source",
]

FILLS = {
    "golden": PatternFill("solid", fgColor="FFFFD700"),
    "blue":   PatternFill("solid", fgColor="FF4285F4"),
    "green":  PatternFill("solid", fgColor="FF00FF00"),
    "yellow": PatternFill("solid", fgColor="FFFFFF00"),
    "red":    PatternFill("solid", fgColor="FFFF0000"),
}
HEADER_FILL = PatternFill("solid", fgColor="FFDDDDDD")


def _norm_url(url: str | None) -> str:
    if not url:
        return ""
    return url.split("?")[0].rstrip("/").lower()


def _norm_name(name: str | None) -> str:
    if not name:
        return ""
    return name.strip().lower()


def _key(name: str | None, url: str | None) -> tuple[str, str]:
    return (_norm_name(name), _norm_url(url))


def _ensure(path: Path, headers: list[str]) -> None:
    if path.exists():
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = path.stem
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    wb.save(path)


def _load(path: Path, headers: list[str]) -> tuple[openpyxl.Workbook, "openpyxl.worksheet.worksheet.Worksheet", list[dict]]:
    _ensure(path, headers)
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    rows: list[dict] = []
    for r in range(2, ws.max_row + 1):
        row = {h: ws.cell(r, c + 1).value for c, h in enumerate(headers)}
        if not row.get("name"):
            continue
        rows.append(row)
    return wb, ws, rows


def _sync_json() -> None:
    """Dump both master xlsx files as JSON siblings for quick diff / verify.
    Runs after every write so json and xlsx never drift."""
    _, _, u = _load(UNTAGGED, UNTAGGED_HEADERS)
    _, _, p = _load(PREDICTIONS, PREDICTIONS_HEADERS)
    UNTAGGED_JSON.write_text(json.dumps(u, indent=2, ensure_ascii=False, default=str))
    PREDICTIONS_JSON.write_text(json.dumps(p, indent=2, ensure_ascii=False, default=str))


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------


def load_untagged() -> list[dict]:
    _, _, rows = _load(UNTAGGED, UNTAGGED_HEADERS)
    return rows


def load_predictions() -> list[dict]:
    _, _, rows = _load(PREDICTIONS, PREDICTIONS_HEADERS)
    return rows


def _load_crocs_keys() -> set[tuple[str, str]]:
    """Load the Crocs sheet (the client's canonical list) for dedupe.
    Crocs rows are NOT duplicated into the master xlsx — they live in
    human_labels.jsonl and we just read keys from them."""
    keys: set[tuple[str, str]] = set()
    if not CROCS_HUMAN_LABELS.exists():
        return keys
    for line in CROCS_HUMAN_LABELS.open():
        r = json.loads(line)
        keys.add(_key(r.get("name"), r.get("linkedin_url")))
    return keys


def exclusion_keys() -> set[tuple[str, str]]:
    """Union of Crocs + untagged + predictions keys. Scraper results
    matching any of these should be skipped before judge runs."""
    keys = _load_crocs_keys()
    for r in load_untagged():
        keys.add(_key(r.get("name"), r.get("linkedin_url")))
    for r in load_predictions():
        keys.add(_key(r.get("name"), r.get("linkedin_url")))
    return keys


def is_known(name: str | None, linkedin_url: str | None) -> bool:
    n, u = _norm_name(name), _norm_url(linkedin_url)
    if not n and not u:
        return False
    for k in exclusion_keys():
        if (u and k[1] == u) or (n and k[0] == n):
            return True
    return False


def add_untagged(
    entries: Iterable[dict],
    source: str,
    seen_at: str | None = None,
) -> int:
    """Append new candidates to untagged.xlsx (dedupe by LinkedIn URL / name).
    Returns the number of newly-added rows."""
    seen_at = seen_at or dt.datetime.now().isoformat(timespec="seconds")
    wb, ws, existing = _load(UNTAGGED, UNTAGGED_HEADERS)
    existing_keys = {_key(r.get("name"), r.get("linkedin_url")) for r in existing}
    added = 0
    for e in entries:
        k = _key(e.get("name"), e.get("linkedin_url"))
        if not k[0] and not k[1]:
            continue
        if k in existing_keys:
            continue
        ws.append([
            e.get("name") or "",
            e.get("linkedin_url") or "",
            e.get("email") or "",
            source,
            seen_at,
        ])
        existing_keys.add(k)
        added += 1
    wb.save(UNTAGGED)
    _sync_json()
    return added


def add_predictions(entries: Iterable[dict], source: str) -> int:
    """Append rated candidates to predictions.xlsx (dedupe). Each row
    also gets its predicted-label cell filled in the label colour.
    Returns the number of newly-added rows."""
    rated_at = dt.datetime.now().isoformat(timespec="seconds")
    wb, ws, existing = _load(PREDICTIONS, PREDICTIONS_HEADERS)
    existing_keys = {_key(r.get("name"), r.get("linkedin_url")) for r in existing}
    added = 0
    for e in entries:
        k = _key(e.get("name"), e.get("linkedin_url"))
        if not k[0] and not k[1]:
            continue
        if k in existing_keys:
            continue
        tags = e.get("reasoning_tags") or []
        if isinstance(tags, list):
            tags = ", ".join(tags)
        ws.append([
            e.get("name") or "",
            e.get("linkedin_url") or "",
            e.get("email") or "",
            e.get("predicted_label") or "",
            tags,
            e.get("reasoning_text") or "",
            e.get("red_bucket") or "",
            e.get("reapproach_after") or "",
            rated_at,
            source,
        ])
        r = ws.max_row
        fill = FILLS.get(e.get("predicted_label", ""))
        if fill:
            # highlight the label column and the name for quick scan
            ws.cell(r, 4).fill = fill
            ws.cell(r, 1).fill = fill
        for c in range(1, len(PREDICTIONS_HEADERS) + 1):
            ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")
        existing_keys.add(k)
        added += 1
    # Column widths
    widths = {1: 24, 2: 40, 3: 28, 4: 14, 5: 40, 6: 60, 7: 22, 8: 18, 9: 22, 10: 28}
    for col, w in widths.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    wb.save(PREDICTIONS)
    _sync_json()
    return added


def stats() -> dict:
    u = load_untagged()
    p = load_predictions()
    return {
        "untagged_count": len(u),
        "predictions_count": len(p),
        "untagged_sources": {r["first_seen_source"]: 0 for r in u}
            and {s: sum(1 for r in u if r["first_seen_source"] == s) for s in {r["first_seen_source"] for r in u if r.get("first_seen_source")}},
        "prediction_labels": {l: sum(1 for r in p if r.get("predicted_label") == l) for l in {r.get("predicted_label") for r in p if r.get("predicted_label")}},
    }


if __name__ == "__main__":
    import json
    print(json.dumps(stats(), indent=2, default=str))
