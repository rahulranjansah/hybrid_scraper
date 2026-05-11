"""
Export judge predictions to an xlsx so the human flagger can review them
with the same color-coding as the source sheet.

Input:  judge/IO/step5a_predict_unflagged/predictions_<lo>_<hi>.jsonl
Output: judge/IO/step5a_predict_unflagged/predictions_<lo>_<hi>.xlsx

Row colors match the predicted label:
  green  -> FF00FF00  (same green as source sheet)
  yellow -> FFFFFF00
  red    -> FFFF0000

Usage:
  uv run python judge/step5a_export_xlsx.py            # rows 68-81 default
  uv run python judge/step5a_export_xlsx.py 100 120    # custom range
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

HERE = Path(__file__).parent
OUT_DIR = HERE / "IO" / "step5a_predict_unflagged"

FILLS = {
    "golden": PatternFill("solid", fgColor="FFFFD700"),  # pure gold
    "blue":   PatternFill("solid", fgColor="FF4285F4"),  # tech-HR
    "green":  PatternFill("solid", fgColor="FF00FF00"),
    "yellow": PatternFill("solid", fgColor="FFFFFF00"),
    "red":    PatternFill("solid", fgColor="FFFF0000"),
}
HEADER_FILL = PatternFill("solid", fgColor="FFDDDDDD")


def main() -> None:
    lo = int(sys.argv[1]) if len(sys.argv) > 1 else 68
    hi = int(sys.argv[2]) if len(sys.argv) > 2 else 81

    src = OUT_DIR / f"predictions_{lo}_{hi}.jsonl"
    dst = OUT_DIR / f"predictions_{lo}_{hi}.xlsx"

    preds = [json.loads(l) for l in src.open()]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Predictions {lo}-{hi}"

    headers = [
        "Sheet Row", "Name", "LinkedIn",
        "Human Flag (if any)", "Judge Prediction",
        "Correct? (fill y/n)", "Reasoning Tags", "Judge Reason",
        "Strengths", "Weaknesses", "Missing Data", "Actionable Insights",
        "Red Bucket", "Reapproach After",
    ]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    def _bullets(items):
        items = items or []
        return "\n".join(f"• {s}" for s in items) if items else ""

    for p in preds:
        label = p.get("predicted_label", "")
        row_vals = [
            p.get("row"),
            p.get("name", ""),
            p.get("linkedin_url", ""),
            p.get("human_flag", "") if p.get("human_flag") != "unflagged" else "",
            label,
            "",  # user fills y/n
            ", ".join(p.get("reasoning_tags", [])),
            p.get("reasoning_text", ""),
            _bullets(p.get("strengths")),
            _bullets(p.get("weaknesses")),
            _bullets(p.get("missing_data")),
            _bullets(p.get("actionable_insights")),
            p.get("red_bucket", "") or "",
            p.get("reapproach_after", "") or "",
        ]
        ws.append(row_vals)
        r = ws.max_row
        # Color the whole data row by the predicted label (like the source sheet).
        fill = FILLS.get(label)
        if fill:
            for c in range(1, len(headers) + 1):
                ws.cell(r, c).fill = fill
        for c in range(1, len(headers) + 1):
            ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")

    # Column widths — tuned so the sheet is readable without resizing.
    widths = {1: 8, 2: 22, 3: 38, 4: 16, 5: 14, 6: 14, 7: 36, 8: 44,
              9: 40, 10: 34, 11: 28, 12: 38, 13: 22, 14: 18}
    for col, w in widths.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    # Freeze header
    ws.freeze_panes = "A2"

    wb.save(dst)
    print(f"Wrote {dst}  ({len(preds)} rows)")


if __name__ == "__main__":
    main()
