"""
Sync master xlsx files to Google Sheets via OAuth.

Sheets to push to are read from .env:
  PREDICTION_FINAL  — edit URL for predictions sheet
  UNTAGGED_DATA     — edit URL for untagged sheet

One-time setup:
  1. console.cloud.google.com -> new project (or pick one)
  2. Enable "Google Sheets API" and "Google Drive API"
  3. Credentials -> Create -> OAuth client ID -> Desktop app
  4. Download JSON, save as:
       judge/IO/master/credentials.json

First run opens a browser for the Google consent screen. The token is
then saved at:
    judge/IO/master/token.json
and reused on subsequent runs.

Usage:
  uv run python judge/sync_sheets.py             # push both sheets
  uv run python judge/sync_sheets.py --only untagged
  uv run python judge/sync_sheets.py --only predictions
"""

from __future__ import annotations

import argparse
import os
import re
import sys
from pathlib import Path

from dotenv import load_dotenv
import gspread
from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request

import openpyxl

HERE = Path(__file__).parent
MASTER_DIR = HERE / "IO" / "master"
CREDENTIALS_FILE = MASTER_DIR / "credentials.json"
TOKEN_FILE = MASTER_DIR / "token.json"
UNTAGGED_XLSX = MASTER_DIR / "untagged.xlsx"
PREDICTIONS_XLSX = MASTER_DIR / "predictions.xlsx"

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive.file",
]


def extract_sheet_id(url: str) -> str:
    """Pull the sheet ID out of a Google Sheets URL."""
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", url)
    if not m:
        raise ValueError(f"couldn't extract sheet id from url: {url!r}")
    return m.group(1)


def get_client() -> gspread.Client:
    if not CREDENTIALS_FILE.exists():
        print(f"ERROR: {CREDENTIALS_FILE} not found.", file=sys.stderr)
        print("Follow setup steps in sync_sheets.py docstring.", file=sys.stderr)
        sys.exit(2)

    creds: Credentials | None = None
    if TOKEN_FILE.exists():
        creds = Credentials.from_authorized_user_file(str(TOKEN_FILE), SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                str(CREDENTIALS_FILE), SCOPES
            )
            # Opens browser for consent.
            creds = flow.run_local_server(port=0)
        TOKEN_FILE.write_text(creds.to_json())
    return gspread.authorize(creds)


def xlsx_to_rows(path: Path) -> list[list]:
    """Read the xlsx into a list of rows (including header). Empty rows dropped."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows: list[list] = []
    for r in range(1, ws.max_row + 1):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if not any(v not in (None, "") for v in row):
            continue
        rows.append([("" if v is None else v) for v in row])
    return rows


def push_to_sheet(client: gspread.Client, url: str, rows: list[list], sheet_name: str) -> str:
    sheet_id = extract_sheet_id(url)
    spreadsheet = client.open_by_key(sheet_id)
    # Use first worksheet; fallback to create one if empty.
    try:
        ws = spreadsheet.get_worksheet(0)
    except Exception:
        ws = spreadsheet.add_worksheet(title=sheet_name, rows=100, cols=20)
    ws.clear()
    if rows:
        ws.update("A1", rows)
    return f"{spreadsheet.title} / {ws.title}"


def main() -> None:
    load_dotenv(HERE.parent / ".env")
    parser = argparse.ArgumentParser(description="Sync master xlsx to Google Sheets")
    parser.add_argument("--only", choices=["untagged", "predictions"],
                        help="sync just one")
    args = parser.parse_args()

    pred_url = os.environ.get("PREDICTION_FINAL", "").strip()
    untag_url = os.environ.get("UNTAGGED_DATA", "").strip()

    if not pred_url and not untag_url:
        print("ERROR: neither PREDICTION_FINAL nor UNTAGGED_DATA set in .env",
              file=sys.stderr)
        sys.exit(2)

    client = get_client()

    if args.only != "predictions" and untag_url:
        rows = xlsx_to_rows(UNTAGGED_XLSX)
        dest = push_to_sheet(client, untag_url, rows, "untagged")
        print(f"  [untagged]    {len(rows)-1} data rows -> {dest}")

    if args.only != "untagged" and pred_url:
        rows = xlsx_to_rows(PREDICTIONS_XLSX)
        dest = push_to_sheet(client, pred_url, rows, "predictions")
        print(f"  [predictions] {len(rows)-1} data rows -> {dest}")


if __name__ == "__main__":
    main()
