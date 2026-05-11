"""
Crocs HR Manager — full end-to-end pipeline.

Stages:
  1. Read the Crocs brief (judge/IO/step3_dspy/crocs_brief.txt).
  2. Derive search keywords from the brief.
  3. combined_scraper stages 1-3.5: query gen -> URL harvest ->
     person extraction -> LinkedIn lookup.
  4. Exclude people already reviewed in human_labels.jsonl.
  5. DSPy judge (combined_scraper.ai_scorer → step3_dspy_judge).
  6. Cap to top N by label priority (golden > green > yellow > red).
  7. Output:
       judge/IO/step5d_pipeline/results_<timestamp>.jsonl
       judge/IO/step5d_pipeline/results_<timestamp>.xlsx (colour-coded)
       judge/IO/step5d_pipeline/results_<timestamp>.md   (via pretty.py)

Usage:
  uv run python judge/run_crocs_hr.py                 # default: top 10
  uv run python judge/run_crocs_hr.py --top 20
  uv run python judge/run_crocs_hr.py --dry-run       # show queries only

Requires GEMINI_API_KEY + BRAVE_API_KEY (+ SERPER_API_KEY for LinkedIn lookup).
"""

from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

HERE = Path(__file__).parent
PROJECT_ROOT = HERE.parent

# Import scraper stages.
sys.path.insert(0, str(PROJECT_ROOT))
from combined_scraper.query_generator import generate_queries, TARGET_SITES  # noqa: E402
from combined_scraper.url_harvester import harvest_urls  # noqa: E402
from combined_scraper.ai_extractor import extract_people  # noqa: E402
from combined_scraper.linkedin_finder import find_linkedin_urls  # noqa: E402
from combined_scraper.ai_scorer import score_results  # noqa: E402 — now DSPy-judge based

BRIEF_FILE = HERE / "IO" / "step3_dspy" / "crocs_brief.txt"
OUT_DIR = HERE / "IO" / "step5d_pipeline"

from master_ledger import (  # noqa: E402
    exclusion_keys, add_untagged, add_predictions, _key,
    UNTAGGED, PREDICTIONS,
)

LABEL_RANK = {"golden": 0, "blue": 1, "green": 2, "yellow": 3, "red": 4, "error": 5}
FILLS = {
    "golden": PatternFill("solid", fgColor="FFFFD700"),
    "blue":   PatternFill("solid", fgColor="FF4285F4"),
    "green":  PatternFill("solid", fgColor="FF00FF00"),
    "yellow": PatternFill("solid", fgColor="FFFFFF00"),
    "red":    PatternFill("solid", fgColor="FFFF0000"),
}
HEADER_FILL = PatternFill("solid", fgColor="FFDDDDDD")


def load_brief() -> str:
    return BRIEF_FILE.read_text()


def brief_to_keywords(brief: str) -> str:
    """Keywords that guide the SCRAPER's query generation. The full brief
    is still passed to the judge separately."""
    return (
        "HR Manager Japan Tokyo, HRBP, HR Business Partner, "
        "People Solutions Manager, Senior HR Manager, HR Director, "
        "Head of HR, consumer goods retail fashion, bilingual English Japanese"
    )


def filter_already_reviewed(results: list[dict]) -> list[dict]:
    """Drop scraper rows whose people are already in the master ledger
    (untagged.xlsx union predictions.xlsx). Saves judge API calls."""
    keys = exclusion_keys()
    if not keys:
        return results
    kept: list[dict] = []
    dropped = 0
    for r in results:
        people = r.get("people") or []
        if not people:
            kept.append(r)
            continue
        hit = False
        for p in people:
            k = _key(p.get("name"), p.get("linkedin_url"))
            for ek in keys:
                if (k[1] and k[1] == ek[1]) or (k[0] and k[0] == ek[0]):
                    hit = True
                    break
            if hit:
                break
        if hit:
            dropped += 1
        else:
            kept.append(r)
    print(f"  [exclude] dropped {dropped} results for already-reviewed people "
          f"(ledger size: {len(keys)})")
    return kept


def take_top_candidates(results: list[dict], top: int) -> list[dict]:
    """Sort by label priority and return top N PERSON-results (not empty rows).
    results is already sorted by score_results; we just cap."""
    person_rows = [
        r for r in results
        if r.get("is_person_result") and r.get("people") and r.get("flag") not in (None, "", "error")
    ]
    person_rows.sort(key=lambda r: LABEL_RANK.get(r.get("flag"), 9))
    return person_rows[:top]


def write_jsonl(rows: list[dict], path: Path) -> None:
    with path.open("w") as f:
        for r in rows:
            f.write(json.dumps(r, ensure_ascii=False, default=str) + "\n")


def write_xlsx(rows: list[dict], path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Crocs HR Shortlist"
    headers = [
        "#", "Name", "Title", "Company", "LinkedIn URL",
        "Judge Label", "Correct? (y/n)",
        "Reasoning Tags", "Judge Reason",
        "Strengths", "Weaknesses", "Missing Data", "Actionable Insights",
        "Source URL", "Red Bucket", "Reapproach After",
    ]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    def _bullets(items: list[str] | None) -> str:
        items = items or []
        if not items:
            return ""
        return "\n".join(f"• {s}" for s in items)

    for i, r in enumerate(rows, 1):
        person = (r.get("people") or [{}])[0]
        label = r.get("flag") or ""
        ws.append([
            i,
            person.get("name") or "",
            person.get("title") or "",
            person.get("company") or "",
            person.get("linkedin_url") or "",
            label,
            "",
            ", ".join(r.get("reasoning_tags") or []),
            r.get("score_reason") or "",
            _bullets(r.get("strengths")),
            _bullets(r.get("weaknesses")),
            _bullets(r.get("missing_data")),
            _bullets(r.get("actionable_insights")),
            r.get("url") or "",
            r.get("red_bucket") or "",
            r.get("reapproach_after") or "",
        ])
        row_idx = ws.max_row
        fill = FILLS.get(label)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row_idx, c)
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = {1: 5, 2: 22, 3: 24, 4: 22, 5: 38, 6: 14, 7: 14,
              8: 36, 9: 48, 10: 42, 11: 36, 12: 30, 13: 40,
              14: 40, 15: 22, 16: 18}
    for col, w in widths.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    ws.freeze_panes = "A2"
    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Crocs HR end-to-end pipeline")
    parser.add_argument("--top", type=int, default=20,
                        help="candidates to keep in the per-run xlsx (default 20)")
    parser.add_argument("--engines", nargs="+", default=["brave", "serper"],
                        choices=["brave", "serper"])
    parser.add_argument("--results-per-query", type=int, default=10,
                        help="results per query per engine (default 10). "
                             "Higher -> more net-new candidates, more cost. "
                             "At 10 expect ~$1 per run and ~10-15 net-new rows.")
    parser.add_argument("--dry-run", action="store_true",
                        help="generate queries only, no search APIs")
    parser.add_argument("--groups", nargs="+", default=None,
                        choices=list(TARGET_SITES.keys()))
    args = parser.parse_args()

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    brief = load_brief()
    keywords = brief_to_keywords(brief)

    print("=" * 60)
    print("CROCS HR — end-to-end pipeline (judge-based)")
    print("=" * 60)
    print(f"Keywords for scraper:\n  {keywords}\n")

    # Stage 1 — query generation
    print("[Stage 1] Generating search queries with Gemini...")
    if args.groups:
        all_queries = []
        for g in args.groups:
            all_queries.extend(generate_queries(keywords, site_group=g))
    else:
        all_queries = generate_queries(keywords)
    print(f"  Generated {len(all_queries)} queries")
    for q in all_queries[:6]:
        print(f"    - [{q.get('site_group')}] {q['query'][:80]}")

    if args.dry_run:
        print("--- DRY RUN — stopping before search ---")
        return

    # Stage 2 — URL harvest
    print(f"\n[Stage 2] Harvesting URLs via {', '.join(args.engines)}...")
    raw = harvest_urls(all_queries, engines=args.engines,
                       results_per_query=args.results_per_query)
    print(f"  Harvested {len(raw)} URLs")

    # Stage 3 — AI person extraction
    print("\n[Stage 3] Extracting people with Gemini...")
    raw = extract_people(raw)

    # Exclude people already reviewed in the human sheet.
    raw = filter_already_reviewed(raw)

    # Stage 3.5 — LinkedIn URL lookup
    print("\n[Stage 3.5] Looking up LinkedIn URLs...")
    raw = find_linkedin_urls(raw)

    # Stage 4 — Judge (the full brief, not just keywords)
    print("\n[Stage 4] Judging with DSPy (4-class rubric)...")
    judged = score_results(raw, keywords_text=keywords, brief=brief)

    # Cap to top N
    top = take_top_candidates(judged, args.top)

    # --- Master ledger updates ---
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    run_source = f"pipeline_{ts}"

    # Every extracted person from this run goes to untagged (the
    # "seen" list). We consider ALL judged person-rows, not just the
    # top-N, so the next pipeline run doesn't re-surface them.
    all_person_rows = [r for r in judged if r.get("is_person_result") and r.get("people")]
    seen_entries = []
    for r in all_person_rows:
        for p in r["people"]:
            seen_entries.append({
                "name": p.get("name") or "",
                "linkedin_url": p.get("linkedin_url") or "",
                "email": "",
            })
    added_seen = add_untagged(seen_entries, source=run_source)
    print(f"  [ledger] +{added_seen} new rows in untagged.xlsx")

    # Only rated candidates go to predictions.xlsx.
    rated_entries = []
    for r in all_person_rows:
        p = r["people"][0]
        rated_entries.append({
            "name": p.get("name") or "",
            "linkedin_url": p.get("linkedin_url") or "",
            "email": "",
            "predicted_label": r.get("flag") or "",
            "reasoning_tags": r.get("reasoning_tags") or [],
            "reasoning_text": r.get("score_reason") or "",
            "red_bucket": r.get("red_bucket") or "",
            "reapproach_after": r.get("reapproach_after") or "",
        })
    added_rated = add_predictions(rated_entries, source=run_source)
    print(f"  [ledger] +{added_rated} new rows in predictions.xlsx")

    # Per-run outputs (timestamped, kept alongside the master ledger)
    jsonl_path = OUT_DIR / f"results_{ts}.jsonl"
    xlsx_path = OUT_DIR / f"results_{ts}.xlsx"
    write_jsonl(top, jsonl_path)
    write_xlsx(top, xlsx_path)

    # Console summary
    print("\n" + "=" * 60)
    print(f"PIPELINE COMPLETE — top {len(top)} of {len(judged)} judged")
    print("=" * 60)
    for i, r in enumerate(top, 1):
        p = (r.get("people") or [{}])[0]
        print(f"  {i:2d}. [{r.get('flag'):7}] {p.get('name','?')} — "
              f"{p.get('title','?')} @ {p.get('company','?')}")
        if p.get("linkedin_url"):
            print(f"        LinkedIn: {p['linkedin_url']}")
    print(f"\nJSONL -> {jsonl_path}")
    print(f"XLSX  -> {xlsx_path}")
    print(f"Ledger -> {UNTAGGED}")
    print(f"Ledger -> {PREDICTIONS}")


if __name__ == "__main__":
    main()
