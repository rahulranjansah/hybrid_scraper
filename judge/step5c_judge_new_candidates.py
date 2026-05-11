"""
Step 5c: Run the judge on a batch of new candidate profiles pasted in
tab-separated form (same column layout as the Crocs sheet: Name /
LinkedIn / Score / Email / Strengths / Weaknesses / Missing Data /
Actionable Insights).

Input:  judge/IO/step5c_new_candidates/raw_input.tsv
Output:
  judge/IO/step5c_new_candidates/predictions.jsonl
  judge/IO/step5c_new_candidates/predictions.md  (via pretty.py)
  judge/IO/step5c_new_candidates/predictions.xlsx

Usage:
  uv run python judge/step5c_judge_new_candidates.py
"""

from __future__ import annotations

import csv
import json
from pathlib import Path

import dspy
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

from step3_dspy_judge import (
    ExplainThenLabel, configure_lm, load_brief, format_candidate,
    derive_label, route_red, reapproach_date, TIMING_CONCERN_TAGS,
)

HERE = Path(__file__).parent
DIR = HERE / "IO" / "step5c_new_candidates"
SRC = DIR / "raw_input.tsv"
JSONL_OUT = DIR / "predictions.jsonl"
XLSX_OUT = DIR / "predictions.xlsx"

FILLS = {
    "golden": PatternFill("solid", fgColor="FFFFD700"),
    "blue":   PatternFill("solid", fgColor="FF4285F4"),
    "green":  PatternFill("solid", fgColor="FF00FF00"),
    "yellow": PatternFill("solid", fgColor="FFFFFF00"),
    "red":    PatternFill("solid", fgColor="FFFF0000"),
}
HEADER_FILL = PatternFill("solid", fgColor="FFDDDDDD")


def split_list(cell: str) -> list[str]:
    if not cell or not cell.strip():
        return []
    return [s.strip() for s in cell.split(",") if s.strip()]


def parse_tsv() -> list[dict]:
    rows: list[dict] = []
    with SRC.open() as f:
        reader = csv.reader(f, delimiter="\t")
        for i, raw in enumerate(reader, start=1):
            raw = raw + [""] * (8 - len(raw))
            name = raw[0].strip()
            if not name:
                continue
            rows.append({
                "index": i,
                "name": name,
                "linkedin_url": raw[1].strip(),
                # Deliberately ignoring the score column (slop).
                "email": raw[3].strip(),
                "strengths": split_list(raw[4]),
                "weaknesses": split_list(raw[5]),
                "missing_data": split_list(raw[6]),
                "actionable_insights": split_list(raw[7]),
                "remark": "",
                "human_flag": "unflagged",
            })
    return rows


def run_judge(rows: list[dict]) -> list[dict]:
    configure_lm()
    brief = load_brief()
    judge = dspy.Predict(ExplainThenLabel)
    preds: list[dict] = []
    for r in rows:
        try:
            p = judge(brief=brief, candidate=format_candidate(r))
            tags = set(p.reasoning_tags or [])
            label = derive_label(tags)
            out = {
                "index": r["index"],
                "name": r["name"],
                "linkedin_url": r.get("linkedin_url", ""),
                "email": r.get("email", ""),
                "predicted_label": label,
                "reasoning_tags": sorted(tags),
                "reasoning_text": p.reasoning_text,
                "strengths": list(p.strengths or []),
                "weaknesses": list(p.weaknesses or []),
                "missing_data": list(p.missing_data or []),
                "actionable_insights": list(p.actionable_insights or []),
            }
            if label == "red":
                bucket = route_red(tags)
                out["red_bucket"] = bucket
                if bucket == "red_reapproach_later":
                    out["reapproach_after"] = reapproach_date()
        except Exception as e:
            out = {"index": r["index"], "name": r["name"], "error": str(e)[:200]}
        preds.append(out)
        print(f"  [{out['index']:2d}] {out.get('predicted_label', 'ERR'):7}  {r['name']}")
    return preds


def write_xlsx(preds: list[dict]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Predictions"
    headers = [
        "#", "Name", "LinkedIn", "Email",
        "Judge Prediction", "Correct? (y/n)",
        "Reasoning Tags", "Judge Reason",
        "Strengths", "Weaknesses", "Missing Data", "Actionable Insights",
        "Red Bucket", "Reapproach After",
    ]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    def _bullets(items):
        items = items or []
        return "\n".join(f"• {s}" for s in items) if items else ""

    for p in preds:
        label = p.get("predicted_label", "")
        ws.append([
            p.get("index"),
            p.get("name", ""),
            p.get("linkedin_url", ""),
            p.get("email", ""),
            label,
            "",
            ", ".join(p.get("reasoning_tags", [])),
            p.get("reasoning_text", ""),
            _bullets(p.get("strengths")),
            _bullets(p.get("weaknesses")),
            _bullets(p.get("missing_data")),
            _bullets(p.get("actionable_insights")),
            p.get("red_bucket", "") or "",
            p.get("reapproach_after", "") or "",
        ])
        r = ws.max_row
        fill = FILLS.get(label)
        if fill:
            for c in range(1, len(headers) + 1):
                ws.cell(r, c).fill = fill
        for c in range(1, len(headers) + 1):
            ws.cell(r, c).alignment = Alignment(wrap_text=True, vertical="top")

    widths = {1: 5, 2: 22, 3: 36, 4: 24, 5: 14, 6: 14, 7: 34, 8: 42,
              9: 40, 10: 34, 11: 28, 12: 38, 13: 22, 14: 18}
    for col, w in widths.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    ws.freeze_panes = "A2"
    wb.save(XLSX_OUT)


def main() -> None:
    rows = parse_tsv()
    print(f"Parsed {len(rows)} candidates from {SRC.name}")
    preds = run_judge(rows)

    with JSONL_OUT.open("w") as f:
        for p in preds:
            f.write(json.dumps(p, ensure_ascii=False) + "\n")
    write_xlsx(preds)

    from collections import Counter
    tally = Counter(p.get("predicted_label") for p in preds if "error" not in p)
    print(f"\nDistribution: {dict(tally)}")
    print(f"JSONL -> {JSONL_OUT}")
    print(f"XLSX  -> {XLSX_OUT}")


if __name__ == "__main__":
    main()
